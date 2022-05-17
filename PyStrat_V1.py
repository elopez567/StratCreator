import tkinter as tk
from tkinter import *
from tkinter import filedialog
import pandas as pd
from win32com.client import Dispatch
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Color, PatternFill, numbers, Border, Side


# Main Application Frame
class Mainapp:
    def __init__(self):
        app = Tk()
        app.title('Strat Creator V1.0')
        app.geometry("1050x450")
        app['background'] = '#003087'
        self.x=2
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.sheet_view.showGridLines = False
        
        # App Labels
        file_label = Label(app, bg='#003087', fg='white', text = "Strat Input File")
        
        ltv_section_label = Label(app, bg='#003087', fg='white', font=("Arial", 15), text = "LTV")
        ltv_start_label = Label(app, bg='#003087', fg='white', text = "Starting Interval")
        ltv_steps_label = Label(app, bg='#003087', fg='white', text = "Steps Up")
        ltv_ceiling_label = Label(app, bg='#003087', fg='white', text = "Ceiling")
        
        fico_section_label = Label(app, bg='#003087', fg='white', font=("Arial", 15), text = "FICO")
        fico_start_label = Label(app, bg='#003087', fg='white', text = "Starting Interval")
        fico_steps_label = Label(app, bg='#003087', fg='white', text = "Steps Up")
        fico_ceiling_label = Label(app, bg='#003087', fg='white', text = "Ceiling")
        
        rate_section_label = Label(app, bg='#003087', fg='white', font=("Arial", 15), text = "Rate")
        rate_start_label = Label(app, bg='#003087', fg='white', text = "Starting Interval")
        rate_steps_label = Label(app, bg='#003087', fg='white', text = "Steps Up")
        rate_ceiling_label = Label(app, bg='#003087', fg='white', text = "Ceiling")
        
        term_section_label = Label(app, bg='#003087', fg='white', font=("Arial", 15), text = "Term")
        term_start_label = Label(app, bg='#003087', fg='white', text = "Starting Interval")
        term_steps_label = Label(app, bg='#003087', fg='white', text = "Steps Up")
        term_ceiling_label = Label(app, bg='#003087', fg='white', text = "Ceiling")
        
        dti_section_label = Label(app, bg='#003087', fg='white', font=("Arial", 15), text = "DTI")
        dti_start_label = Label(app, bg='#003087', fg='white', text = "Starting Interval")
        dti_steps_label = Label(app, bg='#003087', fg='white', text = "Steps Up")
        dti_ceiling_label = Label(app, bg='#003087', fg='white', text = "Ceiling")

        upb_section_label = Label(app, bg='#003087', fg='white', font=("Arial", 15), text = "UPB")
        upb_start_label = Label(app, bg='#003087', fg='white', text = "Starting Interval")
        upb_steps_label = Label(app, bg='#003087', fg='white', text = "Steps Up")
        upb_ceiling_label = Label(app, bg='#003087', fg='white', text = "Ceiling")
        
        age_section_label = Label(app, bg='#003087', fg='white', font=("Arial", 15), text = "Age")
        age_start_label = Label(app, bg='#003087', fg='white', text = "Starting Interval")
        age_steps_label = Label(app, bg='#003087', fg='white', text = "Steps Up")
        age_ceiling_label = Label(app, bg='#003087', fg='white', text = "Ceiling")
        
        # Spaces
        space  = Label(app, bg='#003087', text = "      ")
        space2 = Label(app, bg='#003087', text = "      ")
        space3 = Label(app, bg='#003087', text = "      ")        
        space4 = Label(app, bg='#003087', text = "      ")
        space5 = Label(app, bg='#003087', text = "      ")
        space6 = Label(app, bg='#003087', text = "      ")
        space7 = Label(app, bg='#003087', text = "      ")
        space8 = Label(app, bg='#003087', text = "      ")
        
        # App Variable types
        filedir = StringVar()
        
        self.ltv_starting_int = StringVar()
        self.ltv_starting_int.set('10.0001')
        self.ltv_steps = StringVar()
        self.ltv_steps.set('5.00')
        self.ltv_ceiling = StringVar()
        self.ltv_ceiling.set('95.01')
        
        self.fico_starting_int = StringVar()
        self.fico_starting_int.set('601')
        self.fico_steps = StringVar()
        self.fico_steps.set('50')
        self.fico_ceiling = StringVar()
        self.fico_ceiling.set('851')
        
        self.rate_starting_int = StringVar()
        self.rate_starting_int.set('2.0001')
        self.rate_steps = StringVar()
        self.rate_steps.set('0.250')       
        self.rate_ceiling = StringVar()
        self.rate_ceiling.set('5.01')
        
        self.term_starting_int = StringVar()
        self.term_starting_int.set('121')
        self.term_steps = StringVar()
        self.term_steps.set('60')       
        self.term_ceiling = StringVar()
        self.term_ceiling.set('400')
        
        self.dti_starting_int = StringVar()
        self.dti_starting_int.set('0.0001')
        self.dti_steps = StringVar()
        self.dti_steps.set('5.00')  
        self.dti_ceiling = StringVar()
        self.dti_ceiling.set('65.01')
        
        self.upb_starting_int = StringVar()
        self.upb_starting_int.set('0.001')
        self.upb_steps = StringVar()
        self.upb_steps.set('100000') 
        self.upb_ceiling = StringVar()
        self.upb_ceiling.set('1000000.01') 
        
        self.age_starting_int = StringVar()
        self.age_starting_int.set('0')
        self.age_steps = StringVar()
        self.age_steps.set('7') 
        self.age_ceiling = StringVar()
        self.age_ceiling.set('10') 
        
        
        # App Entries
        file_dir = Entry(app, textvariable=filedir, width=25, borderwidth=3).grid(row=10,column=4, columnspan=5, sticky=tk.W+tk.E)
        
        ltv_start_entry = Entry(app, textvariable=self.ltv_starting_int, width=10, borderwidth=3)
        ltv_step_entry = Entry(app, textvariable=self.ltv_steps, width=10, borderwidth=3)
        ltv_ceiling_entry = Entry(app, textvariable=self.ltv_ceiling, width=10, borderwidth=3)
        
        fico_start_entry = Entry(app, textvariable=self.fico_starting_int, width=10, borderwidth=3)
        fico_step_entry = Entry(app, textvariable=self.fico_steps, width=10, borderwidth=3)   
        fico_ceiling_entry = Entry(app, textvariable=self.fico_ceiling, width=10, borderwidth=3)
            
        rate_start_entry = Entry(app, textvariable=self.rate_starting_int, width=10, borderwidth=3)
        rate_step_entry = Entry(app, textvariable=self.rate_steps, width=10, borderwidth=3)
        rate_ceiling_entry = Entry(app, textvariable=self.rate_ceiling, width=10, borderwidth=3)
        
        term_start_entry = Entry(app, textvariable=self.term_starting_int, width=10, borderwidth=3)
        term_step_entry = Entry(app, textvariable=self.term_steps, width=10, borderwidth=3)        
        term_ceiling_entry = Entry(app, textvariable=self.term_ceiling, width=10, borderwidth=3)
        
        dti_start_entry = Entry(app, textvariable=self.dti_starting_int, width=10, borderwidth=3)
        dti_step_entry = Entry(app, textvariable=self.dti_steps, width=10, borderwidth=3)
        dti_ceiling_entry = Entry(app, textvariable=self.dti_ceiling, width=10, borderwidth=3)
        
        upb_start_entry = Entry(app, textvariable=self.upb_starting_int, width=10, borderwidth=3)
        upb_step_entry = Entry(app, textvariable=self.upb_steps, width=10, borderwidth=3)        
        upb_ceiling_entry = Entry(app, textvariable=self.upb_ceiling, width=10, borderwidth=3)
        
        age_start_entry = Entry(app, textvariable=self.age_starting_int, width=10, borderwidth=3)
        age_step_entry = Entry(app, textvariable=self.age_steps, width=10, borderwidth=3)        
        age_ceiling_entry = Entry(app, textvariable=self.age_ceiling, width=10, borderwidth=3)
        
        # App Grids
        file_label.grid(row=9, column=6)
        
        ltv_section_label.grid(row=2, column=0, padx=10, pady=10)
        ltv_start_label.grid(row=3, column=0, padx=10, pady=10)
        ltv_steps_label.grid(row=5, column=0, padx=10, pady=10)
        ltv_start_entry.grid(row=4, column=0, padx=10, pady=10)
        ltv_step_entry.grid(row=6, column=0, padx=10, pady=10)
        ltv_ceiling_label.grid(row=7, column=0, padx=10, pady=10)
        ltv_ceiling_entry.grid(row=8, column=0, padx=10, pady=10)
        
        space2.grid(row=2, column=1, padx=10, pady=10)
        
        fico_section_label.grid(row=2, column=2, padx=10, pady=10)
        fico_start_label.grid(row=3, column=2, padx=10, pady=10)
        fico_steps_label.grid(row=5, column=2, padx=10, pady=10)
        fico_start_entry.grid(row=4, column=2, padx=10, pady=10)
        fico_step_entry.grid(row=6, column=2, padx=10, pady=10)
        fico_ceiling_label.grid(row=7, column=2, padx=10, pady=10)
        fico_ceiling_entry.grid(row=8, column=2, padx=10, pady=10)
        
        space3.grid(row=2, column=3, padx=10, pady=10)
        
        rate_section_label.grid(row=2, column=4, padx=10, pady=10)
        rate_start_label.grid(row=3, column=4, padx=10, pady=10)
        rate_steps_label.grid(row=5, column=4, padx=10, pady=10)
        rate_start_entry.grid(row=4, column=4, padx=10, pady=10)
        rate_step_entry.grid(row=6, column=4, padx=10, pady=10)
        rate_ceiling_label.grid(row=7, column=4, padx=10, pady=10)
        rate_ceiling_entry.grid(row=8, column=4, padx=10, pady=10)
        
        space4.grid(row=4, column=5, padx=10, pady=10)
        
        term_section_label.grid(row=2, column=6, padx=10, pady=10)
        term_start_label.grid(row=3, column=6, padx=10, pady=10)
        term_steps_label.grid(row=5, column=6, padx=10, pady=10)
        term_start_entry.grid(row=4, column=6, padx=10, pady=10)
        term_step_entry.grid(row=6, column=6, padx=10, pady=10)
        term_ceiling_label.grid(row=7, column=6, padx=10, pady=10)
        term_ceiling_entry.grid(row=8, column=6, padx=10, pady=10)
        
        space5.grid(row=2, column=7, padx=10, pady=10)
        
        dti_section_label.grid(row=2, column=8, padx=10, pady=10)
        dti_start_label.grid(row=3, column=8, padx=10, pady=10)
        dti_steps_label.grid(row=5, column=8, padx=10, pady=10)
        dti_start_entry.grid(row=4, column=8, padx=10, pady=10)
        dti_step_entry.grid(row=6, column=8, padx=10, pady=10)
        dti_ceiling_label.grid(row=7, column=8, padx=10, pady=10)
        dti_ceiling_entry.grid(row=8, column=8, padx=10, pady=10)
        
        space6.grid(row=2, column=9, padx=10, pady=10)
        
        upb_section_label.grid(row=2, column=10)
        upb_start_label.grid(row=3, column=10)
        upb_steps_label.grid(row=5, column=10)
        upb_start_entry.grid(row=4, column=10)
        upb_step_entry.grid(row=6, column=10)
        upb_ceiling_label.grid(row=7, column=10, padx=10, pady=10)
        upb_ceiling_entry.grid(row=8, column=10, padx=10, pady=10)
        
        space7.grid(row=2, column=11, padx=10, pady=10)
        
        age_section_label.grid(row=2, column=12)
        age_start_label.grid(row=3, column=12)
        age_steps_label.grid(row=5, column=12)
        age_start_entry.grid(row=4, column=12)
        age_step_entry.grid(row=6, column=12)
        age_ceiling_label.grid(row=7, column=12, padx=10, pady=10)
        age_ceiling_entry.grid(row=8, column=12, padx=10, pady=10)
        
        space8.grid(row=11,column=4)
        
        
                
        # App Open File Dialog Function
        def openfile():
            app.filename = filedialog.askopenfilename(initialdir="C:/", title="Select Deal Tape")
            self.df = pd.read_excel(io=app.filename)
    
            filedir = StringVar()
            filedir.set(app.filename)
            file_dir = Entry(app, textvariable=filedir, width=25, borderwidth=3).grid(row=10,column=4, columnspan=5, sticky=tk.W+tk.E)
            
        
        # App Create Strat Function
        def runstrat():
            self.x=2
            self.ws.delete_cols(1, 10)
            
            RunField.UPB()
            RunField.Term()
            RunField.FICO()
            RunField.Rate()
            RunField.Age()
            RunField.LTV()
            RunField.DTI()
            RunField.LoanType("Purpose")
            RunField.LoanType("Property")
            RunField.LoanType("Occupancy")
            RunField.LoanType("State")
        
            # Set Column Width
            self.ws.column_dimensions['B'].width = 22
            self.ws.column_dimensions['C'].width = 17
            self.ws.column_dimensions['D'].width = 35
            self.ws.column_dimensions['E'].width = 35
            self.ws.column_dimensions['F'].width = 33
            self.ws.column_dimensions['G'].width = 15
            self.ws.column_dimensions['H'].width = 10
            self.ws.column_dimensions['I'].width = 10
            self.ws.column_dimensions['J'].width = 10
            
            # Save as excel file to same folder as Input file     
            path = app.filename.split('/')
            path.pop()
            outpath =''
            for p in path:
                outpath = outpath + p + '\\'
            outpath = outpath + 'Finished Strat.xlsx'
            self.wb.save(outpath)
            
            # Open finished strat excel
            xl = Dispatch("Excel.Application")
            xl.Visible = 1
            xl.DisplayAlerts = False
            xl.Workbooks.Open(outpath)
        
         # App Buttons
        openfile_btn = Button(app, bg='#31AFDF', fg='white', activebackground='#F1C400', activeforeground='white', command=openfile,text="Open File")
        runstrat_btn = Button(app, bg='#31AFDF', fg='white', activebackground='#F1C400', activeforeground='white', command=runstrat,text="Create Strat")
        openfile_btn.grid(row=12, column=5)
        runstrat_btn.grid(row=12, column=7)
    
        # Collateral Fields & Functions
        class RunField:
            def LTV():
                ltv_bottom = float(self.ltv_starting_int.get())
                ltv_step = float(self.ltv_steps.get())
                ltv_ceiling = float(self.ltv_ceiling.get())
                Format("LTV")
                Grouper(ltv_bottom,ltv_step,ltv_ceiling,"LTV")
                Format("CLTV")
                Grouper(ltv_bottom,ltv_step,ltv_ceiling,"CLTV")
        
            def FICO():
                fico_bottom = float(self.fico_starting_int.get())
                fico_step = float(self.fico_steps.get())
                fico_ceiling = float(self.fico_ceiling.get())
                Format("FICO")
                Grouper(fico_bottom,fico_step,fico_ceiling,"FICO")
                
            def DTI():
                dti_bottom = float(self.dti_starting_int.get())
                dti_step = float(self.dti_steps.get())
                dti_ceiling = float(self.dti_ceiling.get())
                Format("DTI")
                Grouper(dti_bottom,dti_step,dti_ceiling,"DTI")

            def Rate():
                rate_bottom = float(self.rate_starting_int.get())
                rate_step = float(self.rate_steps.get())
                rate_ceiling = float(self.rate_ceiling.get())
                Format("Rate")
                Grouper(rate_bottom,rate_step,rate_ceiling,"Rate")
        
            def UPB():
                upb_bottom = float(self.upb_starting_int.get())
                upb_step = float(self.upb_steps.get())
                upb_ceiling = float(self.upb_ceiling.get())
                Format("UPB")
                Grouper(upb_bottom,upb_step,upb_ceiling,"UPB")
                
            def Term():
                term_bottom = float(self.term_starting_int.get())
                term_step = float(self.term_steps.get())
                term_ceiling = float(self.term_ceiling.get())
                Format("Term")
                Grouper(term_bottom,term_step,term_ceiling,"Term")
                
            def Age():
                age_bottom = float(self.age_starting_int.get())
                age_step = float(self.age_steps.get())
                age_ceiling = float(self.age_ceiling.get())
                Format("Age")
                Grouper(age_bottom,age_step,age_ceiling,"Age")
                
            def LoanType(x):
                Format(x)
                Grouper2(x)                  
                
        # Formatting Function    
        def Format(field):
            
            if field in ['Age','Purpose','Property','Occupancy','State']:
                self.ws[f'B{self.x}'] = f'{field}'
            else:
                self.ws[f'B{self.x}'] = f'Range of \n{field}'
            self.ws[f'C{self.x}'] = 'Number of \nLoans'
            self.ws[f'D{self.x}'] = 'Aggregate Stated \nPrincipal Balance($)'
            self.ws[f'E{self.x}'] = 'Aggregate Stated \nPrincipal Balance(%)' 
            self.ws[f'F{self.x}'] = 'Average Stated \nPrincipal Balance($)'
            self.ws[f'G{self.x}'] = 'WA Note Rate'
            self.ws[f'H{self.x}'] = 'WA FICO'
            self.ws[f'I{self.x}'] = 'WA LTV'
            self.ws[f'J{self.x}'] = 'WA CLTV'
            
            cell_b = self.ws[f'B{self.x}'] 
            cell_c = self.ws[f'C{self.x}'] 
            cell_d = self.ws[f'D{self.x}']
            cell_e = self.ws[f'E{self.x}'] 
            cell_f = self.ws[f'F{self.x}'] 
            cell_g = self.ws[f'G{self.x}'] 
            cell_h = self.ws[f'H{self.x}']
            cell_i = self.ws[f'I{self.x}']
            cell_j = self.ws[f'J{self.x}']
            
            cells = [cell_b, cell_c, cell_d, cell_e, cell_f, cell_g, cell_h, cell_i, cell_j] 
            
            for cell in cells:
                cell.font = Font(bold=True, color="00FFFFFF")
                cell.fill = PatternFill(start_color="003087", end_color="003087", fill_type = "solid")
                
            
            
        # Function that interates through data and groups by desired ranges
        def Grouper(bottom,step,ceiling,field):
            
            self.ws[f'B{self.x-1}'] = field
            self.ws[f'B{self.x-1}'].font = Font(bold=True, size=18)
            while bottom + step < ceiling + step:    
                self.x+=1
                
                top = bottom + step
                self.UPB = self.df["UPB"].loc[(self.df[field] >= bottom) & (self.df[field] < top)]
                self.Rate = self.df["Rate"].loc[(self.df[field] >= bottom) & (self.df[field] < top)]
                self.FICO = self.df["FICO"].loc[(self.df[field] >= bottom) & (self.df[field] < top)]
                self.LTV = self.df["LTV"].loc[(self.df[field] >= bottom) & (self.df[field] < top)]
                self.CLTV = self.df["CLTV"].loc[(self.df[field] >= bottom) & (self.df[field] < top)]

                if field in ["LTV","CLTV","DTI","Rate"]:
                    self.ws[f'B{self.x}'] = f'{round(bottom,2)}1% - {round(top,2)}%'
                elif field == "Rate":
                    self.ws[f'B{self.x}'] = f'{round(bottom,3)}1% - {round(top,3)}%'
                elif field == "UPB":
                    self.ws[f'B{self.x}'] = f'{round(bottom,1):,}1 - {int(top,):,}'
                else:
                    self.ws[f'B{self.x}'] = f'{int(bottom):,} - {int(top-1,):,}'
                
                self.ws[f'C{self.x}'] = self.UPB.count()
                self.ws[f'D{self.x}'] = round(self.UPB.sum(),2)
                self.ws[f'D{self.x}'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                self.ws[f'E{self.x}'] = round((self.UPB.sum() / self.df["UPB"].sum()),4)
                self.ws[f'E{self.x}'].number_format = numbers.FORMAT_PERCENTAGE_00

                if self.UPB.count() == 0:
                    self.ws[f'F{self.x}'] = 0
                    self.ws[f'G{self.x}'] = 0
                    self.ws[f'H{self.x}'] = 0
                    self.ws[f'I{self.x}'] = 0
                    self.ws[f'J{self.x}'] = 0    
                else:
                    self.ws[f'F{self.x}'] = round(self.UPB.sum() / self.UPB.count(),2)
                    self.ws[f'F{self.x}'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    self.ws[f'G{self.x}'] = round(sum(self.UPB*self.Rate)/ self.UPB.sum(),2)/100
                    self.ws[f'G{self.x}'].number_format = numbers.FORMAT_PERCENTAGE_00
                    self.ws[f'H{self.x}'] = int(sum(self.UPB*self.FICO)/ self.UPB.sum())
                    self.ws[f'I{self.x}'] = round(sum(self.UPB*self.LTV)/ self.UPB.sum(),2)
                    self.ws[f'J{self.x}'] = round(sum(self.UPB*self.CLTV)/ self.UPB.sum(),2)    
                bottom = top
            
            # Bottom Border
            thin = Side(border_style="thin", color="000000")
            for row in self.ws[f'B{self.x}:J{self.x}']:
                for cell in row:
                    cell.border = Border(bottom=thin)
  
            
            # Min, Max, Weighted Average
            self.x +=1
            if field != "UPB":
                self.ws[f'B{self.x}'] = f'Min: {round(self.df[field].min(),2)}, Max: {round(self.df[field].max(),2)}, WAvg: {round(sum(self.df["UPB"]*self.df[field])/ self.df["UPB"].sum(),2)}'
            else:
                self.ws[f'B{self.x}'] = f'Average: {round(self.df[field].sum()/self.df[field].count(),2):,}'
            self.ws[f'B{self.x}'].font = Font(bold=True)
            
            self.x+=3
    
       # Function that interates through data and groupsby loan characteristic for fields with no numerical ranges 
        def Grouper2(field):
            self.ws[f'B{self.x-1}'] = field
            self.ws[f'B{self.x-1}'].font = Font(bold=True, size=18)
            
            df2 = self.df.groupby(field, as_index=False)["UPB"].sum()
            df2 = df2.sort_values("UPB", ascending=False)
            
            for i in df2[field]:
                i_upbs = self.df["UPB"].loc[(self.df[field] == i)]
                i_rates = self.df["Rate"].loc[(self.df[field] == i)]
                i_ficos = self.df["FICO"].loc[(self.df[field] == i)]
                i_ltvs = self.df["LTV"].loc[(self.df[field] == i)]
                i_cltvs = self.df["CLTV"].loc[(self.df[field] == i)]
                i_sum = round(self.df["UPB"].loc[(self.df[field] == i)].sum(),2)
                i_count = self.df["UPB"].loc[(self.df[field] == i)].count()
                
                self.x+=1
                self.ws[f'B{self.x}'] = i
                self.ws[f'C{self.x}'] = i_count
                self.ws[f'D{self.x}'] = i_sum
                self.ws[f'D{self.x}'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                self.ws[f'E{self.x}'] = round((i_sum / self.df["UPB"].sum()),4)
                self.ws[f'E{self.x}'].number_format = numbers.FORMAT_PERCENTAGE_00
            
                if i_count == 0:
                    self.ws[f'F{self.x}'] = 0
                    self.ws[f'G{self.x}'] = 0
                    self.ws[f'H{self.x}'] = 0
                    self.ws[f'I{self.x}'] = 0
                    self.ws[f'J{self.x}'] = 0    
                else:
                    self.ws[f'F{self.x}'] = round(i_sum / i_count,2)
                    self.ws[f'F{self.x}'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    self.ws[f'G{self.x}'] = round(sum(i_upbs*i_rates)/ i_sum,2)/100
                    self.ws[f'G{self.x}'].number_format = numbers.FORMAT_PERCENTAGE_00
                    self.ws[f'H{self.x}'] = int(sum(i_upbs*i_ficos)/ i_sum)
                    self.ws[f'I{self.x}'] = round(sum(i_upbs*i_ltvs)/ i_sum,2)
                    self.ws[f'J{self.x}'] = round(sum(i_upbs*i_cltvs)/ i_sum,2)      
            
            # Bottom Border        
            thin = Side(border_style="thin", color="000000")
            for row in self.ws[f'B{self.x}:J{self.x}']:
                for cell in row:
                    cell.border = Border(bottom=thin)
  
            self.x+=3
    
        app.mainloop()

Mainapp()
